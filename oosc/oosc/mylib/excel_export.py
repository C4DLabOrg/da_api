import copy
from sys import stdout

from datetime import datetime
import openpyxl
import os
from datetime import date, timedelta
# from openpyxl.cell import get_column_letter

# wb = openpyxl.load_workbook('juja_road.xlsx')
# print(type(wb))
#
# sheets=wb.get_sheet_names()
# print(sheets)
#
# sheet=""
# if len(sheets)>0:
#     sheet = wb.get_sheet_by_name(sheets[0])
#     print(sheet.title)
from django.core.files.storage import default_storage
from django.utils.dateparse import parse_date
from openpyxl.utils import get_column_letter

from oosc.mylib.common import get_random, get_quick_stream_class_name

months=[{"name":"Sept","days":30},{"name":"Oct","days":31},{"name":"Nov","days":30}]
collumns=["School Name","School Emis Code","Student Id","First Name","Middle name","Last Name","Class Id","Class Name"]


# wb.active.title="Jan"

##Creating the sheets

def excel_generate(queryset):
    wb = openpyxl.Workbook()
    print("generating the excel file")
    number=len(queryset)
    school_name=""
    print("Length ",number)
    maxcols=len(collumns)
    mycollumns=copy.deepcopy(collumns)
    ##append the days to the collumns depending on the number of days
    for j in range(1, months[0]["days"] + 1): mycollumns.append(str(j))

    ##Sheet 1 setting the headers
    sheet=wb.active
    sheet.title=months[0]["name"]
    ## Append the collumn headers to the sheet
    # sheet.column_dimensions[get_column_letter(i+1)].width = 20

    #Setting the headers for the active sheet
    for k,col in enumerate(mycollumns):
        cell = sheet.cell(row=1, column=k + 1)
        cell.value = col
        # cell.style.alignment.wrap_text = True
        # sheet.freeze_panes = 'D1'
        # sheet.freeze_panes = 'E1'
        # sheet.freeze_panes = 'F1'
        # sheet.freeze_panes = 'H1'
        if k+1 <= maxcols:
            sheet.column_dimensions[get_column_letter(k + 1)].width = 20

    #writing data to the sheet
    for i,stud in enumerate(queryset):
        # print()
        # stdout.write( "\r%s of %s" %(str(i+1),str(number)))
        # stdout.flush()
        sheet.cell(row=i+2, column=1).value=stud["school_name"]
        sheet.cell(row=i+2, column=2).value=stud["school_emis_code"]
        sheet.cell(row=i+2, column=3).value=stud["id"]
        sheet.cell(row=i+2, column=4).value=stud["fstname"]
        sheet.cell(row=i+2, column=5).value=stud["midname"]
        sheet.cell(row=i+2, column=6).value=stud["lstname"]
        sheet.cell(row=i+2, column=7).value=stud["class_id"]
        sheet.cell(row=i+2, column=8).value=stud["class_name"]


        ### sheet {i} writing the data
            #setting the headers for the sheet
            # for k,col in enumerate(mycollumns):
            #     cell=sheet.cell(row=1, column=k+1)
            #     cell.value = col
            #     # cell.style.alignment.wrap_text = True
            #     if k+1 <= maxcols:
            #         sheet.column_dimensions[get_column_letter(k + 1)].width = 20
            #
            # #writing the data for sheet {i}
            #         # writing data to the sheet
            # for i, stud in enumerate(queryset):
            #     sheet.cell(row=i + 2, column=1).value = stud["school_name"]
            #     sheet.cell(row=i + 2, column=2).value = stud["school_emis_code"]
            #     sheet.cell(row=i + 2, column=3).value = stud["id"]
            #     sheet.cell(row=i + 2, column=4).value = stud["name"]
            #     sheet.cell(row=i + 2, column=5).value = stud["class_id"]
            #     sheet.cell(row=i + 2, column=6).value = stud["class_name"]
    print("Copy pasting sheets")
    # sheet=wb.create_sheet(index=i, title=month["name"])
    sheet_nov = wb.copy_worksheet(wb.active)
    sheet_nov.title = months[1]["name"]

    sheet_oct = wb.copy_worksheet(wb.active)
    sheet_oct.title = months[2]["name"]

    school_name=queryset[0]["school_name"].replace(" ","_")+"_"+datetime.now().strftime("%d_%b_%Y") if len(queryset) >0 else "oosc_d"
    # print("Done creating the sheets")
    wb.save("oosc.xlsx")
    # print("Saving the file")
    with open("oosc.xlsx") as f:
        default_storage.delete('exports/'+school_name+'.xlsx')
        path = default_storage.save('exports/'+school_name+'.xlsx', f)
    # print (path)
    return path
# wb.save("oosc.xlsx")

def get_age(dob):
    if dob != None:
        days_in_year = 365.2425
        age =int((date.today() - dob).days / days_in_year)
        return "%s"%(age)
    return None

def cal_perc(rw):
    pre=rw["present"]
    total=rw["total_attendance_days"]
    per= float(pre)/float(total)*100
    r="%s %s"%(int(per),"%")
    return r

def export_attendance(queryset,month,year):
    collumns=["County","Subcounty","School Name","School Emis_code",
              "Type","Gender","Age","Class","Guardian","Guardian Phone","Month Days","Days Attendance Taken","Present","Absent","Present %"]
    number = len(queryset)
    school_name = ""
    print("Length ", number)
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Attendance Report %s-%s"%(month,year)
    maxcols = len(collumns)

    ##Setting up the headers
    for k,col in enumerate(collumns):
        cell = sheet.cell(row=1, column=k + 1)
        cell.value = col
        # cell.style.alignment.wrap_text = True
        # sheet.freeze_panes = 'D1'
        # sheet.freeze_panes = 'E1'
        # sheet.freeze_panes = 'F1'
        # sheet.freeze_panes = 'H1'
        if k+1 <= maxcols:
            sheet.column_dimensions[get_column_letter(k + 1)].width = 20

    # writing data to the sheet

    for i, stud in enumerate(queryset):
        # print()
        # stdout.write( "\r%s of %s" %(str(i+1),str(number)))
        # stdout.flush()
        sheet.cell(row=i + 2, column=1).value = stud["county_name"]
        sheet.cell(row=i + 2, column=2).value = stud["subcounty_name"]
        # sheet.cell(row=i + 2, column=3).value = stud["school_name"]
        sheet.cell(row=i + 2, column=4).value = stud["school_emis_code"]
        sheet.cell(row=i + 2, column=5).value = stud["school_type"]
        sheet.cell(row=i + 2, column=6).value = stud["gender"]
        sheet.cell(row=i + 2, column=7).value = get_age(stud["dob_date"])
        sheet.cell(row=i + 2, column=8).value = get_quick_stream_class_name(stud["class_name"])
        sheet.cell(row=i + 2, column=9).value = stud["guardian_name"]
        sheet.cell(row=i + 2, column=10).value = stud["guardian_phone"]
        sheet.cell(row=i + 2, column=11).value = stud["total_month_days"]
        sheet.cell(row=i + 2, column=12).value = stud["total_attendance_days"]
        sheet.cell(row=i + 2, column=13).value = stud["present"]
        sheet.cell(row=i + 2, column=14).value = stud["absent"]
        sheet.cell(row=i + 2, column=15).value = cal_perc(stud)
    print ("the saving name ")
    name=get_random()

    temp_file="%s.xlsx"%(name)
    wb.save(temp_file)
    with open(temp_file) as f:
        fname='exports/%s_%s.xlsx'%(sheet.title,name)
        fname=fname.replace(" ","-")
        default_storage.delete(fname)
        path = default_storage.save(fname, f)

    os.remove(temp_file)
    # print (path)
    return path




