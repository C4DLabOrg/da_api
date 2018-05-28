import openpyxl


# y_name(sheets[0])
#     print(sheet.title)
from django.core.files.storage import default_storage
from openpyxl.utils import get_column_letter


def exportcsv(headers=[], title="Sheet", filename=None, queryset=[]):
    """
    Example
    filename="test"
    queryset=[{"school_title":"Warugara","count":4}]
    headers=[{"name":"School Title","value":"school_title"},{"name":"Students Count","value":"count"}]
    path=exportcsv(filename=filename,queryset=queryset,headers=headers,title="Schools")
    :param headers:
    :param title:
    :param filename:
    :param querset:
    :return:
    """
    path=""
    #####Validate data, assert headers count match an ojbects attributes

    ###Get the totals
    queryset_length = len(queryset)
    headers_length = len(headers)

    ##New workbook
    wb = openpyxl.Workbook()

    ##Create a sheet
    sheet = wb.active
    sheet.title = title

    ###Set the headers

    for k, col in enumerate(headers):
        cell = sheet.cell(row=1, column=k + 1)
        cell.value = col["name"]

        ####Set the size
        if k + 1 <= headers_length:
            sheet.column_dimensions[get_column_letter(k + 1)].width = 20

    ####Writing the data
    for i,data in enumerate(queryset):
        ###Loop through all the headers
        for j,col in enumerate(headers):
            ##i+2 since (i starts at 0, and the row 1 is for headers)
            sheet.cell(row=i+2,column=j+1).value=data[col["value"]]

    ##The output filename
    myfilename="%s.xlsx"%(filename)

    ####Temporatyfilename for openxl
    temp_filename="temp_%s"%(myfilename)

    ##Temporarily save the file
    wb.save(temp_filename)
    with open(temp_filename) as f:
        default_storage.delete('exports/%s'%(myfilename))
        path=default_storage.save('exports/%s'%(myfilename),f)
    return path


if __name__ == '__main__':
    pass

